import tkinter as tk
import tkinter.filedialog
from tkinter import ttk 
import sqlite3
import random
import openpyxl
import datetime


class SHScheduler(tk.Frame): #class inheritance
    '''a program to handle a company's weekle work schedule, the worker's data, etc.'''
    def __init__(self, parentWindow):
        '''
        creates the main window with the main functions
        '''
        tk.Frame.__init__(self, parentWindow) #class inheritance
##        super().__init__(self, parentWindow) #class inheritance
        self.mainWindow = parentWindow
        self.mainWindow.title('Beosztáskezelő')
        print('Initialization...')
        self.loadDatabase('sh_database.db') #if askopenfilename used, some error occurs
        self.listDays()
        self.listShifts()
        self.listWorkers()
        print('Program ready')
        
        year_week = datetime.datetime.now().isocalendar() #isocalendar() method returns a tuple: ISO Year, ISO Week Number, ISO Weekday
        self.year = tk.IntVar()
        self.year.set(year_week[0])
        self.week = tk.IntVar()
        self.week.set(year_week[1])
        self.weekDay = tk.IntVar()
        self.weekDay.set(year_week[2])
        
        tk.Label(self.mainWindow, text='Beosztáskezelő', font=('Helvetica 15 bold')).grid(row=0, column=0)
        tk.Button(self.mainWindow, text='Dolgozók kezelése', width=16, command=self.workerDataManager).grid(row=1, column=0)
        tk.Button(self.mainWindow, text='Munkarend kezelése', width=16, command=self.companyRequestManager).grid(row=2, column=0)
        tk.Button(self.mainWindow, text='Ráérések kezelése', width=16, command=self.workerRequestManager).grid(row=3, column=0)
        tk.Button(self.mainWindow, text='Beosztás kezelése', width=16, command=self.scheduleManager).grid(row=4, column=0)
        tk.Button(self.mainWindow, text='Súgó', width=16, command=self.help).grid(row=5, column=0)
        tk.Button(self.mainWindow, text='Kilépés', width=16, command=self.quit).grid(row=6, column=0)

    def loadDatabase(self, dataBaseFilename=''):
        '''
        loads the database of the given name
        the open file dialog is not working
        '''
        if dataBaseFilename == '':
            self.dataBaseFilename = tk.filedialog.askopenfilename(title='Adatbázis betöltése')
        else:
            self.dataBaseFilename = dataBaseFilename
        self.connection = sqlite3.connect(self.dataBaseFilename)
        self.cursor = self.connection.cursor()
        print('Database: "' + self.dataBaseFilename + '" loaded')

    def listDays(self):
        '''
        lists the days from the database
        '''
        self.cursor.execute('SELECT dayName FROM days')
        arrayDays = self.cursor.fetchall()
        self.days = []
        for i in range(0, len(arrayDays)):
            self.days.append(arrayDays[i][0])
        print('Days listed')

    def listShifts(self):
        '''
        lists the shifts from the database
        '''
        self.cursor.execute('SELECT shiftName FROM shifts ORDER BY shiftId')
        #isActive feature is not working yet
        #self.cursor.execute('SELECT shiftName FROM shifts WHERE isActive = 1 ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        self.shifts = []
        for i in range(0, len(arrayShifts)):
            self.shifts.append(arrayShifts[i][0])
        print('Shifts listed')

    def listWorkers(self):
        '''
        lists the workers from the database sorted by name
        '''
        self.cursor.execute('SELECT workerName FROM workers')
        self.workerNames = []
        workerNamesFetchall = self.cursor.fetchall()
        #print(workerNamesFetchall)
        if workerNamesFetchall != []:
            for row in workerNamesFetchall:
                self.workerNames.append(row[0])
        else:
            self.workerNames.append('')
        self.workerNames.sort()
        #print(self.workerNames)
        print('Workers listed')

    def help(self):
        '''
        opens a help window
        '''
        self.helpWindow = tk.Toplevel()
        self.helpWindow.title('Súgó')
        tk.Label(self.helpWindow, text='Súgó', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        helpText = """Dolgozók kezelése:
    Névválasztó menü: a már az adatbázisban lévő diákok közül lehet választani.
    Dolgozó felvétele: a Név mezőbe beírt névvel új dolgozót vesz fel az adatbázisba.
    Dolgozó törlése: a kiválaszott dolgozót törli az adatbázisból.
    Adatok mentése: az adott nevű dolgozóhoz menti a beírt adatokat.
Munkarend kezelése:
    Kérések kiírása: az adott heti kéréseket jeleníti meg.
    Műszakok kezelése: megadható, mely műszakok aktívak.
    A táblázatban megadható, hogy melyik nap melyik műszakjába hány embert kértek.
    Kérések mentése: az adott heti kéréseket menti az adatbázisba.
Ráérések kezelése:
    Ki kell választani a dolgozó nevét. A táblázatban kipipálható, hogy mely műszakokban ér rá.
    Ráérést lead: elmenti a kiválasztott dolgozó megadott ráéréseit az adott hétre.
Beosztás kezelése:
    Ráérések kiírása: az adott heti ráéréseket jeleníti meg táblázatosan.
    Beosztás készítése: a táblázatban kipipált dolgozókkal elmenti az adott heti beosztást.
    Beosztás kiegészítése: a táblázatban kipipált dolgozókkal készített beosztást automatikusan kiegészíti,
    a beállított algoritmus szerint.
    Algoritmus: a beosztáskészítő által használt algoritmus kiválasztása.
    Beosztás kiírása: az adott hétre már az adatbázisba elmentett beosztást jeleníti meg új ablakban.
    Export xlsx-be: excel táblázatba menti a kiválasztott heti beosztást.
    Beosztás törlése: törli az adott heti beosztást az adatbázisból.
Súgó:
    Megnyitja ezt a súgót.
Kilépés:
    Menti az adatbázist és kilép.
"""
        tk.Label(self.helpWindow, text=helpText, justify='left').grid(row=1, column=0)

    def quit(self):
        '''
        saves the database and closes the program
        '''
        print('Closing...')
        self.saveDatabase()
        self.connection.close()
        self.mainWindow.destroy()

    def saveDatabase(self):
        '''
        saves the database
        '''
        self.connection.commit()
        print('Database saved')
        

#------------------------------------------------------------------------------------------------------
#Worker data
        
    def workerDataManager(self):
        '''
        gui for handling worker data
        '''
        self.workerDataWindow = tk.Toplevel()
        #self.workerDataWindow.grab_set()
        self.workerDataWindow.title('Dolgozók kezelése')
        tk.Label(self.workerDataWindow, text='Dolgozók kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        
        self.workerDataFrame = tk.Frame(self.workerDataWindow, borderwidth=2, relief='ridge')
        self.workerDataFrame.grid(row=1, column=0)
        tk.Label(self.workerDataFrame, text='Név').grid(row=0, column=0)
        self.workerName = tk.StringVar()
        self.workerName.set('név')
        self.nameOptions = ttk.Combobox(self.workerDataFrame, width=18, textvariable=self.workerName)
        self.nameOptions['values'] = self.workerNames
        self.nameOptions.bind('<<ComboboxSelected>>', self.nameMenuSelectionEvent)
        self.nameOptions.grid(row=0, column=1)
        #self.nameOptions.current()
        tk.Button(self.workerDataFrame, text='Dolgozó felvétele', command=self.addWorker).grid(row=0, column=2)
        tk.Button(self.workerDataFrame, text='Dolgozó törlése', command=self.deleteWorker).grid(row=0, column=3)
        tk.Label(self.workerDataFrame, text='Adatok', font=('Helvetica 10 bold')).grid(row=2, column=0, columnspan=2, sticky='W')
        tk.Button(self.workerDataFrame, text='Adatok mentése', command=self.saveWorkerData).grid(row=3, column=2)
        tk.Label(self.workerDataFrame, text='Születési idő').grid(row=3, column=0)
        self.dateOfBirthVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.dateOfBirthVariable).grid(row=3, column=1)
        tk.Label(self.workerDataFrame, text='Telefonszám').grid(row=4, column=0)
        self.phoneNumberVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.phoneNumberVariable).grid(row=4, column=1)
        tk.Label(self.workerDataFrame, text='Tagság érvényessége').grid(row=5, column=0)
        self.membershipValidityVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.membershipValidityVariable).grid(row=5, column=1)
        tk.Label(self.workerDataFrame, text='Aktív').grid(row=6, column=0)
        self.isActiveVariable = tk.BooleanVar()
        self.isActiveCheckbutton = tk.Checkbutton(self.workerDataFrame, variable=self.isActiveVariable)
        self.isActiveCheckbutton.grid(row=6, column=1)

    def nameMenuSelectionEvent(self, event):
        '''
        this function is called when you select a name from the dropdown list
        it loads the data of the selected worker
        '''
        workerName = self.workerName.get()
        self.cursor.execute('SELECT dateOfBirth FROM workers WHERE workerName = ?', (workerName, ))
        self.dateOfBirthVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT phoneNumber FROM workers WHERE workerName = ?', (workerName, ))
        self.phoneNumberVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT membershipValidity FROM workers WHERE workerName = ?', (workerName, ))
        self.membershipValidityVariable.set( self.cursor.fetchone()[0] )
        #isActive feature is not working yet
        #self.isActiveVariable.set( True if self.cursor.fetchone()[0] == 1 else False )
        self.cursor.execute('SELECT isActive FROM workers WHERE workerName = ?', (workerName, ))
        if self.cursor.fetchone()[0] == 1:
            self.isActiveCheckbutton.select()
        else:
            self.isActiveCheckbutton.deselect()

    def addWorker(self):
        '''
        adds the worker with the given name to the database
        calls saveWorkerData() to save the other data for the worker
        '''
        workerName = self.workerName.get()
        if workerName != '':
            self.saveWorkerData()
            self.listWorkers()
            self.nameOptions.destroy() #destroy and recreate widget
            self.nameOptions = ttk.Combobox(self.workerDataFrame, width=18, textvariable=self.workerName)
            self.nameOptions['values'] = self.workerNames
            self.nameOptions.bind('<<ComboboxSelected>>', self.nameMenuSelectionEvent)
            self.nameOptions.grid(row=0, column=1)
        print(workerName + ' added')

    def saveWorkerData(self):
        '''
        saves data (birthday, phone number, etc.) for the worker
        '''
        workerName = self.workerName.get()
        dateOfBirth = self.dateOfBirthVariable.get()
        phoneNumber = self.phoneNumberVariable.get()
        membershipValidity = self.membershipValidityVariable.get()
        isActive = self.isActiveVariable.get()
        try: #if the worker is not in the database, insert
            self.cursor.execute('INSERT INTO workers (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive) VALUES (?, ?, ?, ?, ?)', (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive))
        except: #if the worker is already in the database, update
            self.cursor.execute('UPDATE workers SET dateOfBirth = "' + dateOfBirth + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET phoneNumber = "' + phoneNumber + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET membershipValidity = "' + membershipValidity + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET isActive = "' + str(int(isActive)) + '" WHERE workerName = "' + workerName + '"')
        self.saveDatabase()
        
    def deleteWorker(self):
        '''
        deletes the worker with the given name from the database
        '''
        workerName = self.workerName.get()
        self.cursor.execute('DELETE FROM workers WHERE workerName = ?', (workerName, ))
        self.saveDatabase()
        self.listWorkers()
        self.nameOptions.destroy() #destroy and recreate widget
        self.nameOptions = ttk.Combobox(self.workerDataFrame, width=18, textvariable=self.workerName)
        self.nameOptions['values'] = self.workerNames
        self.nameOptions.bind('<<ComboboxSelected>>', self.nameMenuSelectionEvent)
        self.nameOptions.grid(row=0, column=1)
        print(workerName + ' deleted')


#------------------------------------------------------------------------------------------------------
#Company requests
    
    def companyRequestManager(self):
        '''
        gui for handling company requests
        '''
        self.companyRequestWindow = tk.Toplevel()
        #self.companyRequestWindow.grab_set()
        self.companyRequestWindow.title('Munkarend kezelése')
        tk.Label(self.companyRequestWindow, text='Munkarend kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

        #miscFrame
        self.miscFrame = tk.Frame(self.companyRequestWindow, borderwidth=2, relief='ridge')
        self.miscFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
        tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
        tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
        tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
        tk.Button(self.miscFrame, text='Kérések kiírása', command=self.showCompanyRequestFrame).grid(row=0, column=4)
        tk.Button(self.miscFrame, text='Műszakok kezelése', command=self.shiftManager).grid(row=1, column=0, columnspan=2)
        tk.Button(self.miscFrame, text='Ráérések kezelése', command=self.workerRequestManager).grid(row=2, column=0, columnspan=2)

        #companyRequestFrame
        self.showCompanyRequestFrame()

    def showCompanyRequestFrame(self):
        try:
            self.companyRequestFrame.destroy()
        except:
            pass
        self.companyRequestFrame = tk.Frame(self.companyRequestWindow, borderwidth=2, relief='ridge')
        self.companyRequestFrame.grid(row=2, column=0, sticky='W')
        tk.Button(self.companyRequestFrame, text='Kérések mentése', command=self.saveCompanyRequest).grid(row=1, column=1, columnspan=2)
        #create the field of entries
        date = datetime.datetime.fromisocalendar(self.year.get(), self.week.get(), self.weekDay.get())
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.companyRequestFrame, text=text, width=8).grid(row=2, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.companyRequestFrame, text=self.shifts[i], width=8).grid(row=3+i, column=0)
        self.companyRequestEntries, self.companyRequestVariables = [], [] #lists to store the entries and their variables
        for j in range(0, len(self.days)):
            self.companyRequestEntries.append([])
            self.companyRequestVariables.append([])
            for i in range(0, len(self.shifts)):
                variable = tk.IntVar()
                entry = tk.Entry(self.companyRequestFrame, textvariable=variable, width=5)
                entry.grid(row=3+i, column=1+j)
                self.companyRequestEntries[j].append(entry)
                self.companyRequestVariables[j].append(variable)
        #load the previously saved company request
        #and fill the field of entries with the data
        self.loadAndShowCompanyRequest()

    def createCompanyRequest(self):
        '''
        creates a table for company requests for the given week
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequest_' + str(year) + '_' + str(week) + ' AS SELECT * FROM companyRequest WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.cursor.execute('INSERT OR IGNORE INTO companyRequest_' + str(year) + '_' + str(week) +
                                    ' (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
                                    (dayId, shiftId, workerNumber) )
        self.saveDatabase()

    def loadAndShowCompanyRequest(self):
        '''
        loads company requests for the given week
        and fills the previousley created entry table with the data
        '''
        self.createCompanyRequest()
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                isActive = self.cursor.fetchone()[0]
                if isActive == 1:
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    try:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    except:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    self.companyRequestVariables[j][i].set(workerNumber)

    def getCompanyRequest(self):
        '''
        takes the numbers from the entry table into a numpy array
        '''
        self.companyRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.companyRequestGrid[i][j] = self.companyRequestVariables[j][i].get()
        #print(self.companyRequestGrid)
        
    def saveCompanyRequest(self):
        '''
        saves company requests for the given week to the database
        first calls getCompanyRequest() in order to get the data from the entry field
        '''
        self.getCompanyRequest()
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute('INSERT OR IGNORE INTO companyRequest (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
##                                    (dayId, shiftId, int(self.companyRequestGrid[i][j])) ) #cast a numpy value to int: value.item()
                self.cursor.execute('UPDATE companyRequest_' + str(year) + '_' + str(week) + 
                                    ' SET workerNumber = ' + str(int(self.companyRequestGrid[i][j])) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #update és insert egyszerre: ha a meglévő érték nem azonos a beírttal, frissíteni kell
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Company requests - Shift manager

    def shiftManager(self):
        '''
        gui for managing shifts
        '''
        self.shiftManagerWindow = tk.Toplevel()
        #self.shiftManagerWindow.grab_set()
        self.shiftManagerWindow.title('Műszakok kezelése')
        tk.Label(self.shiftManagerWindow, text='Műszakok kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.miscFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
        self.miscFrame.grid(row=1, column=0, sticky='W')
        tk.Button(self.miscFrame, text='Új műszak', command=self.addShiftManager).grid(row=0, column=0)
        tk.Button(self.miscFrame, text='Műszakok mentése', command=self.saveShifts).grid(row=0, column=1)
        self.shiftsFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
        self.shiftsFrame.grid(row=2, column=0, sticky='W')
        self.shiftCheckbuttons, self.shiftVariables = [], []
        for i in range(0, len(self.shifts)):
            tk.Label(self.shiftsFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
            self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
            isActive = self.cursor.fetchone()[0]
            variable = tk.BooleanVar()
            variable.set(isActive)
            checkbutton = tk.Checkbutton(self.shiftsFrame, variable=variable)
            checkbutton.grid(row=2+i, column=1)
            self.shiftCheckbuttons.append(checkbutton)
            self.shiftVariables.append(variable)
        
    def addShiftManager(self):
        '''
        gui for adding new shifts
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        self.addShiftWindow = tk.Toplevel()
        #self.addShiftWindow.grab_set()
        self.addShiftWindow.title('Új műszak')
        tk.Label(self.addShiftWindow, text='Új műszak felvétele', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.addShiftFrame = tk.Frame(self.addShiftWindow, borderwidth=2, relief='ridge')
        self.addShiftFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.addShiftFrame, text='Műszak neve').grid(row=0, column=0)
        self.newShiftName = tk.StringVar()
        tk.Entry(self.addShiftFrame, textvariable=self.newShiftName).grid(row=0, column=1)
        tk.Button(self.addShiftFrame, text='Műszak felvétele', command=self.addNewShift).grid(row=1, column=0)
        
    def addNewShift(self):
        '''
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        newShiftName = self.newShiftName.get()
        self.cursor.execute('INSERT INTO shifts (shiftName, isActive) VALUES (?, ?)', (newShiftName, 1, ))
        #update the list of shifts
        self.saveDatabase()

    def saveShifts(self):
        '''
        saves the shifts table in the database
        '''
        for i in range(0, len(self.shifts)):
            shiftName = self.shifts[i]
            isActive = self.shiftVariables[i].get()
            isActive = 1 if isActive == True else 0
            self.cursor.execute('UPDATE shifts SET isActive = "' + str(isActive) + '" WHERE shiftName = "' + shiftName + '"')
        self.saveDatabase()

#------------------------------------------------------------------------------------------------------
#Worker requests

    def workerRequestManager(self):
        '''
        gui for handling worker requests
        '''
        tableExists = 1
        year = self.year.get()
        week = self.week.get()
        try:
            self.cursor.execute('SELECT * FROM companyRequest_' + str(year) + '_' + str(week))
        except:
            tableExists = 0
            
        if tableExists == 0:
            text = 'Table workerRequests_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
            self.messageWindow = tk.Toplevel()
            self.messageWindow.grab_set()
            tk.Label(self.messageWindow, text=text).grid(row=0, column=0)
        else:
            self.workerRequestWindow = tk.Toplevel()
            #self.workerRequestWindow.grab_set()
            self.workerRequestWindow.title('Ráérések kezelése')
            tk.Label(self.workerRequestWindow, text='Ráérések kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

            self.miscFrame = tk.Frame(self.workerRequestWindow, borderwidth=2, relief='ridge')
            self.miscFrame.grid(row=1, column=0, sticky='W')
            tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
            tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
            tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
            tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
            tk.Button(self.miscFrame, text='Ráéréstábla kirajzolása', command=self.showWorkerRequestGrid).grid(row=0, column=4)
            tk.Label(self.miscFrame, text='Név').grid(row=2, column=0)
            self.workerName = tk.StringVar()
            self.workerName.set('név')
            self.nameOptions = ttk.Combobox(self.miscFrame, width=20, textvariable=self.workerName)
            self.nameOptions['values'] = self.workerNames
            self.nameOptions.bind('<<ComboboxSelected>>', self.optionMenuSelectionEvent)
            self.nameOptions.grid(row=2, column=1, columnspan=4)
            tk.Button(self.miscFrame, text='Ráérést lead', command=self.saveWorkerRequest).grid(row=3, column=1)
            tk.Button(self.miscFrame, text='Beosztás kezelése', command=self.scheduleManager).grid(row=4, column=1)
            
            self.workerRequestFrame = tk.Frame(self.workerRequestWindow, borderwidth=2, relief='ridge')
            self.workerRequestFrame.grid(row=2, column=0, sticky='W')
            self.showWorkerRequestGrid()

    def showWorkerRequestGrid(self):
        '''
        shows a check grid for selecting the requests for the given worker
        '''
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.workerRequestFrame, text=text, width=8).grid(row=1, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.workerRequestFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
        self.requestCheckbuttons, self.requestVariables = [], [] #lists to store the entries and their variables
        for j in range(0, len(self.days)):
            self.requestCheckbuttons.append([])
            self.requestVariables.append([])
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                     ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #a checkbutton should be active only if
                #the requested number of workers is greater than 0 for the given shift
                if self.cursor.fetchone()[0] > 0:
                    variable = tk.BooleanVar()
                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
                    checkbutton.grid(row=2+i, column=1+j)
                    self.requestCheckbuttons[j].append(checkbutton)
                    self.requestVariables[j].append(variable)
                else:
                    variable = tk.BooleanVar()
                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
                    checkbutton.grid(row=2+i, column=1+j)
                    checkbutton['state'] = 'disabled'
                    self.requestCheckbuttons[j].append(checkbutton)
                    self.requestVariables[j].append(variable)
        #print(self.requestVariables)

    def optionMenuSelectionEvent(self, event):
        '''
        event for selecting a name
        first it deselects all checkbuttons
        then it checks the shifts the worker requested for the given week
        '''
        for daysCheckbuttons in self.requestCheckbuttons:
            for checkbutton in daysCheckbuttons:
                checkbutton.deselect()
        year = self.year.get()
        week = self.week.get()
        workerName = self.workerName.get()
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerIds = [row[0] for row in self.cursor.fetchall()]
                    if workerId in workerIds:
                        self.requestCheckbuttons[j][i].select()
                except:
                    pass

    def getWorkerRequest(self):
        '''
        takes the checks from the check table into a numpy array (1 if checked, else 0)
        '''
        workerName = self.workerName.get()
        self.workerRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.workerRequestGrid[i][j] = 1 if self.requestVariables[j][i].get() else 0 #when creating these checkbuttons and variables, the indices are reversed
        #print(workerName, '\n', self.workerRequestGrid)

    def saveWorkerRequest(self):
        '''
        saves worker requests for the given week to the database 
        '''
        self.getWorkerRequest()
        workerName = self.workerName.get()
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId, dayId, shiftId, UNIQUE(workerId, dayId, shiftId))')
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                if self.workerRequestGrid[i][j] == 1:
                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
                                        ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, j, i))
                else:
                    try:
                        self.cursor.execute('DELETE FROM workerRequests_' + str(year) + '_' + str(week) +
                                            ' WHERE workerId = ' + str(workerId) + ' AND dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    except:
                        pass
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Schedule manager

    def scheduleManager(self):
        '''
        gui for creating schedule
        '''
        tableExists = 1
        year = self.year.get()
        week = self.week.get()
        try:
            self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week))
        except:
            tableExists = 0
            
        if tableExists == 0:
            text = 'Table workerRequests_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
            self.messageWindow = tk.Toplevel()
            self.messageWindow.grab_set()
            tk.Label(self.messageWindow, text=text).grid(row=0, column=0)
        else:
            self.scheduleWindow = tk.Toplevel()
            #self.scheduleWindow.grab_set()
            tk.Label(self.scheduleWindow, text='Beosztás kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

            self.miscFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
            self.miscFrame.grid(row=1, column=0, sticky='W')
            tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
            tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
            tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
            tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
            tk.Button(self.miscFrame, text='Ráérések kiírása', command=self.showWorkerRequests).grid(row=0, column=4, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás készítése', command=self.createSchedule).grid(row=1, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás kiegészítése', command=self.fillCreatedSchedule).grid(row=1, column=2, columnspan=2)
            tk.Label(self.miscFrame, text='Algoritmus').grid(row=1, column=4)
            self.algorithmList = ['random', 'frommin']
            self.algorithmVar = tk.StringVar()
            self.algorithmVar.set(self.algorithmList[0])
            tk.OptionMenu(self.miscFrame, self.algorithmVar, *self.algorithmList).grid(row=1, column=5)
            tk.Button(self.miscFrame, text='Beosztás kiírása', command=self.showSchedule).grid(row=2, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Export xlsx-be', command=self.scheduleExportXlsx).grid(row=2, column=2, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás törlése', command=self.deleteSchedule).grid(row=3, column=0, columnspan=2)
            self.showWorkerRequests()

    def loadRequestsListToShow(self, table):
        '''
        loads worker max number of request for shifts for the week into a list
        '''
        requests = list(range(len(self.shifts))) #[0]*len(self.shifts)
        year = self.year.get()
        week = self.week.get()
        tableExists = 1
        if table == 'workerRequests':
            try:
                self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week))
            except:
                tableExists = 0
        elif table == 'companyRequest':
            try:
                self.cursor.execute('SELECT workerNumber FROM ' + table + '_' + str(year) + '_' + str(week))
            except:
                tableExists = 0
            
        if tableExists == 1:
            for j in range(0, len(self.days)):
                self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
                dayId = self.cursor.fetchone()[0]
                for i in range(0, len(self.shifts)):
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    if table == 'workerRequests':
                        self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week) +
                                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                        workerIds = self.cursor.fetchall()
                        if len(workerIds) >= requests[i]:
                            requests[i] = len(workerIds)
                    elif table == 'companyRequest':
                        self.cursor.execute('SELECT workerNumber FROM ' + table + '_' + str(year) + '_' + str(week) +
                                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                        workerNumber = self.cursor.fetchone()[0]
                        if workerNumber >= requests[i]:
                            requests[i] = workerNumber
            #print(table, requests)
        else:
            text = 'Table ' + table + '_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
        return requests

    def showWorkerRequests(self):
        '''
        creates a frame for handling worker requests for the given week
        checks if a worker is scheduled
        '''
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        row = 1 #same as gridRow
        requests = self.loadRequestsListToShow('workerRequests')
##        requestsSum = sum(requests)
##        if requestsSum > 0:
        try:
            self.scheduleFrame.destroy() #if exists
        except:
            pass
        self.scheduleFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
        self.scheduleFrame.grid(row=2, column=0, sticky='W')
        self.scheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.scheduleFrame))
        self.scheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.scheduleFrame))
        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        tk.Label(self.scheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.scheduleFrame, text=text, width=12, font='Helvetica 10 bold').grid(row=0, column=1+2*j, columnspan=2) #!!!!!!!!! column(span)
        for i in range(0, len(self.shifts)):
            tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            self.scheduleByHandCheckbuttons.append([])
            self.scheduleByHandVariables.append([])
            self.scheduleByHandNameLabels.append([])
            gridRow = 1 #same as row
            gridRow_ = gridRow #to track the last empty row
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.scheduleByHandCheckbuttons[j].append([])
                self.scheduleByHandVariables[j].append([])
                self.scheduleByHandNameLabels[j].append([])
                tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=gridRow, column=0)
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = self.cursor.fetchall()
                for k in range(0, requests[i]):
                    try:
                        workerId = workerIds[k][0]
                        self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + str(workerId))
                        workerName = self.cursor.fetchone()[0]
                        nameLabel = tk.Label(self.scheduleFrame, text=workerName)
                        nameLabel.grid(row=gridRow_, column=1+2*j) #!!!!!!!!! column
                        self.scheduleByHandNameLabels[j][i].append(nameLabel)
                        variable = tk.BooleanVar()
                        checkbutton = tk.Checkbutton(self.scheduleFrame, variable=variable, command=lambda x1=j, x2=i, x3=k, x4=workerName: self.disableWorkerSelection(x1, x2, x3, x4))
                        checkbutton.grid(row=gridRow_, column=1+2*j+1) #!!!!!!!!! column
                        try:
                            #check if the worker to be shown is already scheduled there (in a previous run of the program)
                            self.cursor.execute( 'SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                                 ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                            if workerId in [ workerIds[0] for workerIds in self.cursor.fetchall()]:
                                #if a worker is scheduled, check the box
                                checkbutton.select()
                        except:
                            pass
                        self.scheduleByHandCheckbuttons[j][i].append(checkbutton)
                        self.scheduleByHandVariables[j][i].append([variable, workerId, workerName])
                    except:
                        #shitty solution to fill empty spaces (rowconfigure?)
                        tk.Label(self.scheduleFrame, text='').grid(row=gridRow_, column=1+j)
                    gridRow_ += 1
                gridRow = gridRow + requests[i]
        #print(self.scheduleByHandVariables)

    def loadSchedule(self):
        '''
        loads the schedule and the backups for the given week
        '''
        year = self.year.get()
        week = self.week.get()
        self.schedule = [] #
        self.backup = [] #
        for j in range(0, len(self.days)):
            self.schedule.append([]) #
            self.backup.append([]) #
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.schedule[j].append(workerNames)
                
                #load the backup workers for the week (same as loading the scheduled workers)
                self.cursor.execute('SELECT workerId FROM backup_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.backup[j].append(workerNames)
        #print(self.schedule)

    def showSchedule(self):
        '''
        loads the schedule for the given week
        and shows it in a seperate window
        '''
        year = self.year.get()
        week = self.week.get()
        try:
            self.loadSchedule()
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleFrame = tk.Frame(self.showScheduleWindow, borderwidth=2, relief='ridge')
            self.showScheduleFrame.grid(row=3, column=0, sticky='W')
            
            self.showScheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.showScheduleFrame))
            self.showScheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.showScheduleFrame))
            
            #requests = [4, 1, 4]
            requests = self.loadRequestsListToShow('companyRequest')
            #print(requests)
            row = 1 #starting row is the one under the day names
            tk.Label(self.showScheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
            for j in range(0, len(self.days)):
                tk.Label(self.showScheduleFrame, text=self.days[j], width=12, font='Helvetica 10 bold').grid(row=0, column=1+j)
            for i in range(0, len(self.shifts)):
                tk.Label(self.showScheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
                row = row + requests[i]
            for j in range(0, len(self.days)):
                row_ = 1
                row = row_
                for i in range(0, len(self.shifts)):
                    for k in range(0, requests[i]):
                        try:
                            workerName = self.schedule[j][i][k]
                        except:
                            workerName = ''
                        tk.Label(self.showScheduleFrame, text=workerName).grid(row=row, column=1+j)
                        row += 1
                row_ = row_ + requests[i]
        except:
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleWindow.grab_set()
            text = 'Table schedule_' + str(year) + '_' + str(week) + ' does not exist.'
            tk.Label(self.showScheduleWindow, text=text).grid(row=0, column=0)

    def scheduleExportXlsx(self):
        '''
        exports the schedule for the given week into a .xlsx file
        first loads the schedule from the database
        saves the backup workers for the week on a different worksheet (same as loading the scheduled workers)
        '''
        self.loadSchedule()
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        filename = 'schedule_' + str(year) + '_' + str(week) + '.xlsx'
        #schedule
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'schedule_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution below
        requests = self.loadRequestsListToShow('companyRequest')
        row = 2
        worksheet.cell(row=1, column=1).value = str(year) + '/' + str(week)
        worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            worksheet.cell(row=1, column=2+j).value = text
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.schedule[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        #backup
        worksheet = workbook.create_sheet()
        worksheet.title = 'backup_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution above
        row = 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.backup[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        workbook.save(filename=filename)
        print('Schedule exported')

    def deleteSchedule(self):
        '''
        deletes schedule for the given week
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_' + str(year) + '_' + str(week))
        self.cursor.execute('DROP TABLE IF EXISTS backup_' + str(year) + '_' + str(week))
        self.saveDatabase()
    
    def disableWorkerSelection(self, column, row, row_k, nameToDisable):
        '''
        if someone is scheduled to work in a shfit, he/she can't work in another shift on the given day
        the possibility to check him/her into another shift is disabled
        '''
        #print(column, row, row_k, nameToDisable)
        #print(self.scheduleByHandVariables[column][row][row_k][0].get(), self.scheduleByHandNameLabels[column][row][row_k]['text'])
        if self.scheduleByHandVariables[column][row][row_k][0].get() == True:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'disabled'
        else:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'normal'

        self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[column], ))
        dayId = self.cursor.fetchone()[0]
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[row], ))
        shiftId = self.cursor.fetchone()[0]
        self.disableWorkerSelectionForShift(dayId, shiftId, column, row)
        
    def disableWorkerSelectionForShift(self, dayId, shiftId, column, row):
        '''
        if the workers requested by the company for a given shift is met,
        the possibility to check other workers for that shift is disabled
        '''
        year = self.year.get()
        week = self.week.get()
        requests = self.loadRequestsListToShow('workerRequests') #gives the max number of requests for shifts
        workersScheduledForShift = []
        workerNumberScheduled = 0
        workersScheduledForDay = []
        for k in range(0, requests[row]):
            #try-except is not the most elegant solution
            #it is for overcoming that requests list contains the max number of requests for shifts (for example [8, 1, 5])
            #and the real requests for a given day can be fewer (for example [6, 1, 4])
            #so the index k may result in out of range error
            #and company requests is [4, 1, 4]
            try:
                if self.scheduleByHandVariables[column][row][k][0].get() == True:
                    workerNumberScheduled += 1
                    workersScheduledForShift.append(self.scheduleByHandNameLabels[column][row][k]['text'])
            except:
                pass
        #print('workersScheduledForShift: ', workersScheduledForShift)

        for row_ in range(0, len(requests)):
            for k in range(0, requests[row_]):
                try:
                    if self.scheduleByHandVariables[column][row_][k][0].get() == True:
                        workersScheduledForDay.append(self.scheduleByHandNameLabels[column][row_][k]['text'])
                except:
                    pass
        #print('workersScheduledForDay: ', workersScheduledForDay)

        self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) + 
                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
        workerNumber = self.cursor.fetchone()[0]
        if workerNumberScheduled == workerNumber:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'disabled'
            except:
                pass
        else:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'normal'
            except:
                pass
            
    def highlightOn(self, event, frame):
        '''
        when the mouse hovers over a name, highlights all of his/her requests for the week in red
        '''
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='red')
        except:
            pass

    def highlightOff(self, event, frame):
        '''
        disables highlighting defined in highlightOn()
        '''
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='black')
        except:
            pass

    def createBackup(self):
        '''
        creates a backup table for the given week
        from the workers who are not scheduled
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS backup_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE backup_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) )
                workerIdsSchedule = self.cursor.fetchall()
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIdsRequests = self.cursor.fetchall()
                for id_ in workerIdsRequests:
                    if not id_ in workerIdsSchedule:
                        row = id_[0], dayId, shiftId
                        self.cursor.execute('INSERT OR IGNORE INTO backup_' + str(year) + '_' + str(week) +
                                            ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', row)
        print('Backup created')

    def createSchedule(self):
        '''
        creates schedule from the check table
        also calls createBackup()
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE schedule_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for day in range(0, len(self.scheduleByHandVariables)):
            for shift in range(0, len(self.scheduleByHandVariables[day])):
                for row in self.scheduleByHandVariables[day][shift]:
                    if row[0].get()==True:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', (row[1], day, shift) )
        self.createBackup()
        self.saveDatabase()
        print('Schedule created')


    def fillCreatedSchedule(self):
        '''
        completes and saves the schedule from the check table based on the selected algorithm
        also calls createBackup() and then getNumberOfScheduledDays()
        '''
        year = self.year.get()
        week = self.week.get()
        self.createSchedule()
        self.cursor.execute('DROP TABLE IF EXISTS companyRequestModified') #this modified table is for counting how many workers are still needed 
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequestModified AS SELECT * FROM companyRequest_'
                            + str(year) + '_' + str(week) +' WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                partialScheduledWorkers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNeeded = self.cursor.fetchone()[0]
                if workerNeeded > len(partialScheduledWorkers):
                    workerNeeded -= len(partialScheduledWorkers)
                else:
                    workerNeeded = 0
                self.cursor.execute('INSERT INTO companyRequestModified (dayId, shiftId, workerNumber) VALUES (?, ?, ?) ', (dayId, shiftId, workerNeeded))

        if self.algorithmVar.get() == 'random':
            #completes the schedule randomly, may not schedule workers who are free anyway
            self.cursor.execute( 'SELECT * FROM workerRequests_' + str(year) + '_' + str(week) )
            workerRequests = self.cursor.fetchall()
            random.shuffle(workerRequests)
            for row in workerRequests:
                workerId, dayId, shiftId = row[0], row[1], row[2]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + #select workerId instead of *
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workersNeeded = self.cursor.fetchone()[0]
                if len(workers) < workersNeeded:
                    if not workerId in workers:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        elif self.algorithmVar.get() == 'frommin':
            #completes the table starting from the workers who requested the least days
            self.getNumberOfRequestedDays()
            self.cursor.execute('SELECT * FROM workers ORDER BY requestedDaysWeekly')
            workers = self.cursor.fetchall()
            for worker in workers:
                workerId = worker[0]
                self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE workerId = ' + str(workerId) )
                for row in self.cursor.fetchall():
                    dayId, shiftId = row[1], row[2]
                    self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + 
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workers = self.cursor.fetchall()
                    self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workersNeeded = self.cursor.fetchone()[0]
                    if len(workers) < workersNeeded:
                        if not workerId in workers:
                            self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                                '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        self.createBackup()
        self.getNumberOfScheduledDays()
        self.saveDatabase()
        print('Schedule created and filled')

    def getNumberOfRequestedDays(self):
        '''
        determines how many days each worker has requested for the week
        '''
        self.numberOfRequestedDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM workerRequests_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfRequestedDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET requestedDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfRequestedDays:', self.numberOfRequestedDays)

    def getNumberOfScheduledDays(self):
        '''
        determines how many days each worker has been scheduled for for the week
        '''
        self.numberOfScheduledDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM schedule_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfScheduledDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET scheduledDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfScheduledDays: ' + self.numberOfScheduledDays)



        
if __name__ == '__main__':
    root = tk.Tk()
    SH = SHScheduler(root)
    SH.grid(row=0, column=0)
    #print('class __doc__:', SH.__doc__)
    root.mainloop()
