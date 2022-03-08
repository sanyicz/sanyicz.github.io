from PyQt5 import QtWidgets
from PyQt5 import QtCore
from PyQt5 import QtGui
import sys
import sqlite3
import random
import openpyxl
import datetime


class Label(QtWidgets.QLabel):
    def __init__(self, text, layout):
        QtWidgets.QLabel.__init__(self, text) #or super().init__(self, text)
        self.layout = layout
        self.workerName = text

    def enterEvent(self, event):
        index = self.layout.count()
        while(index >= 0):
            try:
                nameLabel = self.layout.itemAt(index).widget()
                if nameLabel.text() == self.workerName and nameLabel.text() != '':
                    nameLabel.setStyleSheet('color: red')
            except:
                pass
            index -=1

    def leaveEvent(self, event):
        index = self.layout.count()
        while(index >= 0):
            try:
                nameLabel = self.layout.itemAt(index).widget()
                if nameLabel.text() == self.workerName and nameLabel.text() != '':
                    nameLabel.setStyleSheet('color: black')
            except:
                pass
            index -=1

        
class SHScheduler(QtWidgets.QApplication):
    '''a program to handle a company's weekle work schedule, the worker's data, etc.'''
    def __init__(self, args):
        '''
        creates the main window with the main functions
        '''
        #is it ok to inherit like this?
        #or: no inheritance, delete this and uncomment rows in main
        super().__init__(args) #or super(SHScheduler, self).__init__(args) -> why?
        
        self.mainWindow = QtWidgets.QWidget()
        self.mainWindow.setWindowTitle('Beosztáskezelő')
        
        print('Initialization...')
        self.loadDatabase('sh_database.db') #if askopenfilename used, some error occurs
        self.listDays()
        self.listShifts()
        self.listWorkers()
        print('Program ready')
        
        year_week = datetime.datetime.now().isocalendar() #isocalendar() method returns a tuple: ISO Year, ISO Week Number, ISO Weekday
        self.year = year_week[0]
        self.week = year_week[1]
        self.weekDay = year_week[2]
        
        self.label1 = QtWidgets.QLabel('Beosztáskezelő')
        self.button1 = QtWidgets.QPushButton('Dolgozók kezelése')
        self.button1.clicked.connect(self.workerDataManager)
        self.button2 = QtWidgets.QPushButton('Munkarend kezelése')
        self.button2.clicked.connect(self.companyRequestManager)
        self.button3 = QtWidgets.QPushButton('Ráérések kezelése')
        self.button3.clicked.connect(self.workerRequestManager)
        self.button4 = QtWidgets.QPushButton('Beosztás kezelése')
        self.button4.clicked.connect(self.scheduleManager)
        self.button5 = QtWidgets.QPushButton('Súgó')
        self.button5.clicked.connect(self.help)
        self.button6 = QtWidgets.QPushButton('Kilépés')
        self.button6.clicked.connect(self.quit)
        
        self.mainLayout = QtWidgets.QVBoxLayout()
        self.mainLayout.addWidget(self.label1)
        self.mainLayout.addWidget(self.button1)
        self.mainLayout.addWidget(self.button2)
        self.mainLayout.addWidget(self.button3)
        self.mainLayout.addWidget(self.button4)
        self.mainLayout.addWidget(self.button5)
        self.mainLayout.addWidget(self.button6)
        
        self.mainWindow.setLayout(self.mainLayout)
        self.mainWindow.show()

    def setYear(self, year):
        self.year = year
        #print(self.year)

    def setWeek(self, week):
        self.week = week
        #print(self.week)
        
    def loadDatabase(self, dataBaseFilename=''):
        '''
        loads the database of the given name
        the open file dialog is not working
        '''
        if dataBaseFilename == '':
            self.dataBaseFilename = tk.filedialog.askopenfilename(title='Adatbázis betöltése')
        else:
            self.dataBaseFilename = dataBaseFilename
        self.connection = sqlite3.connect(self.dataBaseFilename)
        self.cursor = self.connection.cursor()
        print('Database: "' + self.dataBaseFilename + '" loaded')

    def listDays(self):
        '''
        lists the days from the database
        '''
        self.cursor.execute('SELECT dayName FROM days')
        arrayDays = self.cursor.fetchall()
        self.days = []
        for i in range(0, len(arrayDays)):
            self.days.append(arrayDays[i][0])
        print('Days listed')

    def listShifts(self):
        '''
        lists the shifts from the database
        '''
        self.cursor.execute('SELECT shiftName FROM shifts ORDER BY shiftId')
        #isActive feature is not working yet
        #self.cursor.execute('SELECT shiftName FROM shifts WHERE isActive = 1 ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        self.shifts = []
        for i in range(0, len(arrayShifts)):
            self.shifts.append(arrayShifts[i][0])
        print('Shifts listed')

    def listWorkers(self):
        '''
        lists the workers from the database sorted by name
        '''
        self.cursor.execute('SELECT workerName FROM workers')
        self.workerNames = []
        workerNamesFetchall = self.cursor.fetchall()
        #print(workerNamesFetchall)
        if workerNamesFetchall != []:
            for row in workerNamesFetchall:
                self.workerNames.append(row[0])
        else:
            self.workerNames.append('')
        self.workerNames.sort()
        #print(self.workerNames)
        print('Workers listed')

    def help(self):
        '''
        opens a help window
        '''
        self.helpWindow = QtWidgets.QWidget()
        self.helpWindow.setWindowTitle('Súgó')
        helpLabel = QtWidgets.QLabel(self.helpWindow)
        helpText = "Dolgozók kezelése:\nNévválasztó menü: a már az adatbázisban lévő diákok közül lehet választani.\nDolgozó felvétele: a Név mezőbe beírt névvel új dolgozót vesz fel az adatbázisba\nDolgozó törlése: a kiválaszott dolgozót törli az adatbázisból.\nAdatok mentése: az adott nevű dolgozóhoz menti a beírt adatokat."
        helpLabel.setText(helpText)
        #helpLabel.setWordWrap(True)
        self.helpWindow.show()

    def quit(self):
        '''
        saves the database and closes the program
        '''
        print('Closing...')
        self.saveDatabase()
        self.connection.close()
        self.mainWindow.destroy()
        self.exit(0) #same as calling QCoreApplication.quit(), but it's overloaded by this quit method

    def saveDatabase(self):
        '''
        saves the database
        '''
        self.connection.commit()
        print('Database saved')
        

#------------------------------------------------------------------------------------------------------
#Worker data
        
    def workerDataManager(self):
        '''
        gui for handling worker data
        '''
        self.workerDataWindow = QtWidgets.QWidget()
        self.workerDataWindow.setWindowTitle('Dolgozók kezelése')
        layout = QtWidgets.QGridLayout()

        headerLabel = QtWidgets.QLabel('Dolgozók kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        nameLabel = QtWidgets.QLabel('Név')
        self.nameOptions = QtWidgets.QComboBox()
        self.nameOptions.addItems(self.workerNames)
##        self.nameOptions.currentTextChanged.connect(self.nameMenuSelectionEvent) #???????????
        self.nameOptions.activated.connect(self.nameMenuSelectionEvent)
        self.workerNameVariable = QtWidgets.QLineEdit()
        addWorkerButton = QtWidgets.QPushButton('Dolgozó felvétele')
        addWorkerButton.clicked.connect(self.addWorker)
        delWorkerButton = QtWidgets.QPushButton('Dolgozó törlése')
        delWorkerButton.clicked.connect(self.deleteWorker)
        dataLabel = QtWidgets.QLabel('Adatok')
        dateOfBirthLabel = QtWidgets.QLabel('Születési idő')
        self.dateOfBirthVariable = QtWidgets.QLineEdit()
        saveDataButton = QtWidgets.QPushButton('Adatok mentése')
        saveDataButton.clicked.connect(self.saveWorkerData)
        phoneNumberLabel = QtWidgets.QLabel('Telefonszám')
        self.phoneNumberVariable = QtWidgets.QLineEdit()
        membershipLabel = QtWidgets.QLabel('Tagság érvényessége')
        self.membershipValidityVariable = QtWidgets.QLineEdit()
        isActiveLabel = QtWidgets.QLabel('Aktív')
        self.isActiveVariable = QtWidgets.QCheckBox()
        
        miscLayout = QtWidgets.QGridLayout()
        miscLayout.addWidget(nameLabel, 0, 0)
        miscLayout.addWidget(self.nameOptions, 0, 1)
        miscLayout.addWidget(self.workerNameVariable, 1, 1)
        miscLayout.addWidget(addWorkerButton, 1, 2)
        miscLayout.addWidget(delWorkerButton, 1, 3)
        miscLayout.addWidget(dataLabel, 2, 0)
        miscLayout.addWidget(dateOfBirthLabel, 3, 0)
        miscLayout.addWidget(self.dateOfBirthVariable, 3, 1)
        miscLayout.addWidget(saveDataButton, 3, 2)
        miscLayout.addWidget(phoneNumberLabel, 4, 0)
        miscLayout.addWidget(self.phoneNumberVariable, 4, 1)
        miscLayout.addWidget(membershipLabel, 5, 0)
        miscLayout.addWidget(self.membershipValidityVariable, 5, 1)
        miscLayout.addWidget(isActiveLabel, 6, 0)
        miscLayout.addWidget(self.isActiveVariable, 6, 1)
        layout.addLayout(miscLayout, 1, 0)

        self.workerDataWindow.setLayout(layout)
        self.workerDataWindow.show()

    def nameMenuSelectionEvent(self, event):
        '''
        this function is called when you select a name from the dropdown list
        it loads the data of the selected worker
        '''
        workerName = self.nameOptions.currentText()
        self.workerNameVariable.setText(workerName)
        self.cursor.execute('SELECT dateOfBirth FROM workers WHERE workerName = ?', (workerName, ))
        self.dateOfBirthVariable.setText( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT phoneNumber FROM workers WHERE workerName = ?', (workerName, ))
        self.phoneNumberVariable.setText( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT membershipValidity FROM workers WHERE workerName = ?', (workerName, ))
        self.membershipValidityVariable.setText( self.cursor.fetchone()[0] )
        #isActive feature is not working yet
        #(self.isActiveVariable.setChecked() if self.cursor.fetchone()[0] == 1)
        self.cursor.execute('SELECT isActive FROM workers WHERE workerName = ?', (workerName, ))
        if self.cursor.fetchone()[0] == 1:
            self.isActiveVariable.setChecked(True)
        else:
            self.isActiveVariable.setChecked(False)

    def addWorker(self):
        '''
        adds the worker with the given name to the database
        calls saveWorkerData() to save the other data for the worker
        '''
        workerName = self.workerNameVariable.text()
        if workerName != '':
            self.saveWorkerData()
            self.listWorkers()
            self.nameOptions.clear()
            self.nameOptions.addItems(self.workerNames)
        print(workerName + ' added')

    def saveWorkerData(self):
        '''
        saves data (birthday, phone number, etc.) for the worker
        '''
        workerName = self.workerNameVariable.text()
        dateOfBirth = self.dateOfBirthVariable.text()
        phoneNumber = self.phoneNumberVariable.text()
        membershipValidity = self.membershipValidityVariable.text()
        isActive = self.isActiveVariable.isChecked()
        try: #if the worker is not in the database, insert
            self.cursor.execute('INSERT INTO workers (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive) VALUES (?, ?, ?, ?, ?)', (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive))
        except: #if the worker is already in the database, update
            self.cursor.execute('UPDATE workers SET dateOfBirth = "' + dateOfBirth + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET phoneNumber = "' + phoneNumber + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET membershipValidity = "' + membershipValidity + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET isActive = "' + str(int(isActive)) + '" WHERE workerName = "' + workerName + '"')
        self.saveDatabase()
        
    def deleteWorker(self):
        '''
        deletes the worker with the given name from the database
        '''
        workerName = self.workerNameVariable.text()
        self.cursor.execute('DELETE FROM workers WHERE workerName = ?', (workerName, ))
        self.saveDatabase()
        self.listWorkers()
        self.nameOptions.clear()
        self.nameOptions.addItems(self.workerNames)
        print(workerName + ' deleted')


#------------------------------------------------------------------------------------------------------
#Company requests
    
    def companyRequestManager(self):
        '''
        gui for handling company requests
        '''
        self.companyRequestWindow = QtWidgets.QWidget()
        self.companyRequestWindow.setWindowTitle('Munkarend kezelése')
        layout = QtWidgets.QGridLayout()

        #header
        headerLabel = QtWidgets.QLabel('Munkarend kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        #miscFrame
        yearLabel = QtWidgets.QLabel('Év')
        yearEntry = QtWidgets.QLineEdit()
        yearEntry.textChanged.connect(lambda x=yearEntry.text(): self.setYear(x))
        yearEntry.setText(str(self.year))
        weekLabel = QtWidgets.QLabel('Hét')
        weekEntry = QtWidgets.QLineEdit()
        weekEntry.textChanged.connect(lambda x=weekEntry.text(): self.setWeek(x))
        weekEntry.setText(str(self.week))
        showButton = QtWidgets.QPushButton('Kérések kiírása')
        showButton.clicked.connect(self.showCompanyRequestFrame)
        shiftsButton = QtWidgets.QPushButton('Műszakok kezelése')
        shiftsButton.clicked.connect(self.shiftManager)
        requestsButton = QtWidgets.QPushButton('Ráérések kezelése')
        requestsButton.clicked.connect(self.workerRequestManager)
        miscLayout = QtWidgets.QGridLayout() #layout
        miscLayout.addWidget(yearLabel, 0, 0)
        miscLayout.addWidget(yearEntry, 0, 1)
        miscLayout.addWidget(weekLabel, 0, 2)
        miscLayout.addWidget(weekEntry, 0, 3)
        miscLayout.addWidget(showButton, 0, 4)
        miscLayout.addWidget(shiftsButton, 1, 0)
        miscLayout.addWidget(requestsButton, 2, 0)
        layout.addLayout(miscLayout, 1, 0)

        #companyRequestFrame
        self.companyRequestLayout = QtWidgets.QGridLayout() #layout
        layout.addLayout(self.companyRequestLayout, 2, 0)
        
        self.companyRequestWindow.setLayout(layout)
        self.companyRequestWindow.show()
        
        self.showCompanyRequestFrame()

    def showCompanyRequestFrame(self):
        try:
            while self.companyRequestLayout.count():
                widget = self.companyRequestLayout.itemAt(0).widget()
                self.companyRequestLayout.removeWidget(widget)
                widget.setParent(None)
        except:
            pass
        saveRequestsButton = QtWidgets.QPushButton('Kérések mentése')
        saveRequestsButton.clicked.connect(self.saveCompanyRequest)
        self.companyRequestLayout.addWidget(saveRequestsButton, 1, 1, 1, 2) #positon (1,1), occupies 1 row and 2 columns
        #create the field of entries
        year = self.year
        week = self.week
        weekDay = self.weekDay
        date = datetime.datetime.fromisocalendar(int(year), int(week), int(weekDay))
        startDate = date - datetime.timedelta(days=self.weekDay-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            label = QtWidgets.QLabel(text)
            self.companyRequestLayout.addWidget(label, 2, 1+j)
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            self.companyRequestLayout.addWidget(label, 3+i, 0)
        self.companyRequestEntries = [] #lists to store the entries and their variables
        self.companyRequestVariables = [] #????
        for j in range(0, len(self.days)):
            self.companyRequestEntries.append([])
            self.companyRequestVariables.append([])
            for i in range(0, len(self.shifts)):
                entry = QtWidgets.QLineEdit()
                self.companyRequestLayout.addWidget(entry, 3+i, 1+j)
                self.companyRequestEntries[j].append(entry)
                #self.companyRequestVariables[j].append(variable) #????

        #load the previously saved company request
        #and fill the field of entries with the data
        self.loadAndShowCompanyRequest()
        
    def createCompanyRequest(self):
        '''
        creates a table for company requests for the given week
        '''
        year = self.year
        week = self.week
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequest_' + str(year) + '_' + str(week) + ' AS SELECT * FROM companyRequest WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.cursor.execute('INSERT OR IGNORE INTO companyRequest_' + str(year) + '_' + str(week) +
                                    ' (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
                                    (dayId, shiftId, workerNumber) )
        self.saveDatabase()

    def loadAndShowCompanyRequest(self):
        '''
        loads company requests for the given week
        and fills the previousley created entry table with the data
        '''
        self.createCompanyRequest()
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                isActive = self.cursor.fetchone()[0]
                if isActive == 1:
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    try:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    except:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    self.companyRequestEntries[j][i].setText(str(workerNumber))

    def getCompanyRequest(self):
        '''
        takes the numbers from the entry table into a numpy array
        '''
        self.companyRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.companyRequestGrid[i][j] = int(self.companyRequestEntries[j][i].text())
        #print(self.companyRequestGrid)
        
    def saveCompanyRequest(self):
        '''
        saves company requests for the given week to the database
        first calls getCompanyRequest() in order to get the data from the entry field
        '''
        self.getCompanyRequest()
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute('INSERT OR IGNORE INTO companyRequest (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
##                                    (dayId, shiftId, int(self.companyRequestGrid[i][j])) ) #cast a numpy value to int: value.item()
                self.cursor.execute('UPDATE companyRequest_' + str(year) + '_' + str(week) + 
                                    ' SET workerNumber = ' + str(int(self.companyRequestGrid[i][j])) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #update és insert egyszerre: ha a meglévő érték nem azonos a beírttal, frissíteni kell
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Company requests - Shift manager

    def shiftManager(self):
        '''
        gui for managing shifts
        '''
        self.shiftManagerWindow = QtWidgets.QWidget()
        self.shiftManagerWindow.setWindowTitle('Műszakok kezelése')
        layout = QtWidgets.QGridLayout()

        #header
        headerLabel = QtWidgets.QLabel('Műszakok kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)
        
        #miscFrame
        addShiftButton = QtWidgets.QPushButton('Új műszak')
        addShiftButton.clicked.connect(self.addShiftManager)
        saveShiftsButton = QtWidgets.QPushButton('Műszakok mentése')
        saveShiftsButton.clicked.connect(self.saveShifts)
        miscLayout = QtWidgets.QGridLayout() #layout
        miscLayout.addWidget(addShiftButton, 0, 0)
        miscLayout.addWidget(saveShiftsButton, 0, 1)
        layout.addLayout(miscLayout, 1, 0)

        #shiftsFrame
        self.shiftCheckbuttons = []
        self.shiftVariables = [] #????
        shiftsLayout = QtWidgets.QGridLayout() #layout
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            shiftsLayout.addWidget(label, 2+i, 0)
            self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
            isActive = self.cursor.fetchone()[0]
            checkbutton = QtWidgets.QCheckBox()
            checkbutton.setChecked(True if isActive else False)
            shiftsLayout.addWidget(checkbutton, 2+i, 1)
            self.shiftCheckbuttons.append(checkbutton)
        layout.addLayout(shiftsLayout, 2, 0)

        self.shiftManagerWindow.setLayout(layout)
        self.shiftManagerWindow.show()
        
    def addShiftManager(self):
        '''
        gui for adding new shifts
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        self.addShiftWindow = QtWidgets.QWidget()
        self.addShiftWindow.setWindowTitle('Új műszak')
        layout = QtWidgets.QGridLayout()

        #header
        headerLabel = QtWidgets.QLabel('Új műszak felvétele')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        #addShiftFrame
        shiftNameLabel = QtWidgets.QLabel('Műszak neve')
        self.shiftNameEntry = QtWidgets.QLineEdit()
        addShiftButton = QtWidgets.QPushButton('Műszak felvétele')
        addShiftButton.clicked.connect(self.addNewShift)
        addShiftLayout = QtWidgets.QGridLayout()
        addShiftLayout.addWidget(shiftNameLabel, 0, 0)
        addShiftLayout.addWidget(self.shiftNameEntry, 0, 1)
        addShiftLayout.addWidget(addShiftButton, 1, 0)
        layout.addLayout(addShiftLayout, 1, 0)

        self.addShiftWindow.setLayout(layout)
        self.addShiftWindow.show()
        
    def addNewShift(self):
        '''
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        newShiftName = self.shiftNameEntry.text()
        self.cursor.execute('INSERT INTO shifts (shiftName, isActive) VALUES (?, ?)', (newShiftName, 1, ))
        #need to be done: update the list of shifts
        self.saveDatabase()

    def saveShifts(self):
        '''
        saves the shifts table in the database
        '''
        for i in range(0, len(self.shifts)):
            shiftName = self.shifts[i]
##            isActive = self.shiftCheckbuttons[i].isChecked()
##            isActive = 1 if isActive == True else 0
            isActive = 1 if self.shiftCheckbuttons[i].isChecked() == True else 0
            self.cursor.execute('UPDATE shifts SET isActive = "' + str(isActive) + '" WHERE shiftName = "' + shiftName + '"')
        self.saveDatabase()

#------------------------------------------------------------------------------------------------------
#Worker requests

    def workerRequestManager(self):
        '''
        gui for handling worker requests
        '''
        tableExists = 1
        year = self.year
        week = self.week
        try:
            self.cursor.execute('SELECT * FROM companyRequest_' + str(year) + '_' + str(week))
        except:
            tableExists = 0

        if tableExists == 0:
            text = 'Table companyRequest_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
            self.messageWindow = QtWidgets.QWidget()
            layout = QtWidgets.QVBoxLayout() #layout
            label = QtWidgets.QLabel(text)
            layout.addWidget(label)
            self.messageWindow.setLayout(layout)
            self.messageWindow.show()
        else:
            self.workerRequestWindow = QtWidgets.QWidget()
            self.workerRequestWindow.setWindowTitle('Ráérések kezelése')
            self.workerRequestManagerlayout = QtWidgets.QGridLayout()

            #header
            headerLabel = QtWidgets.QLabel('Ráérések kezelése')
            headerLayout = QtWidgets.QVBoxLayout()
            headerLayout.addWidget(headerLabel)
            self.workerRequestManagerlayout.addLayout(headerLayout, 0, 0)

            #miscFrame
            yearLabel = QtWidgets.QLabel('Év')
            yearEntry = QtWidgets.QLineEdit()
            yearEntry.textChanged.connect(lambda x=yearEntry.text(): self.setYear(x))
            yearEntry.setText(str(self.year))
            weekLabel = QtWidgets.QLabel('Hét')
            weekEntry = QtWidgets.QLineEdit()
            weekEntry.textChanged.connect(lambda x=weekEntry.text(): self.setWeek(x))
            weekEntry.setText(str(self.week))
            weeklyRequestButton = QtWidgets.QPushButton('Ráéréstábla kirajzolása')
            weeklyRequestButton.clicked.connect(self.showWorkerRequestGrid)
            nameLabel = QtWidgets.QLabel('Név')
            self.nameOptions = QtWidgets.QComboBox()
            self.nameOptions.addItems(self.workerNames)
            self.nameOptions.activated.connect(self.optionMenuSelectionEvent)
            saveButton = QtWidgets.QPushButton('Ráérést lead')
            saveButton.clicked.connect(self.saveWorkerRequest)
            scheduleButton = QtWidgets.QPushButton('Beosztás kezelése')
            scheduleButton.clicked.connect(self.scheduleManager)
            miscLayout = QtWidgets.QGridLayout() #layout
            miscLayout.addWidget(yearLabel, 0, 0)
            miscLayout.addWidget(yearEntry, 0, 1)
            miscLayout.addWidget(weekLabel, 0, 2)
            miscLayout.addWidget(weekEntry, 0, 3)
            miscLayout.addWidget(weeklyRequestButton, 0, 4)
            miscLayout.addWidget(nameLabel, 2, 0)
            miscLayout.addWidget(self.nameOptions, 2, 1, 1, 4)
            miscLayout.addWidget(saveButton, 3, 1)
            miscLayout.addWidget(scheduleButton, 4, 1)
            self.workerRequestManagerlayout.addLayout(miscLayout, 1, 0)

            #workerRequestFrame
            self.workerRequestLayout = QtWidgets.QGridLayout() #layout
            self.workerRequestManagerlayout.addLayout(self.workerRequestLayout, 2, 0)
            
            self.workerRequestWindow.setLayout(self.workerRequestManagerlayout)
            self.workerRequestWindow.show()

            self.showWorkerRequestGrid()

    def showWorkerRequestGrid(self):
        '''
        shows a check grid for selecting the requests for the given worker
        '''
        year = self.year
        week = self.week
        weekDay = self.weekDay
        date = datetime.datetime.fromisocalendar(int(year), int(week), int(weekDay))
        startDate = date - datetime.timedelta(days=self.weekDay-1)
        #endDate = startDate + datetime.timedelta(days=6)
        try:
            while self.workerRequestLayout.count():
                widget = self.workerRequestLayout.itemAt(0).widget()
                self.workerRequestLayout.removeWidget(widget)
                widget.setParent(None)
        except:
            pass
        
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            label = QtWidgets.QLabel(text)
            self.workerRequestLayout.addWidget(label, 1, 1+j)
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            self.workerRequestLayout.addWidget(label, 2+i, 0)
        self.requestCheckbuttons = [] #lists to store the entries and their variables
        self.requestVariables = [] #????
        #if companyRequest_year_week doesn't exists, load companyRequest_year_week-1
        for j in range(0, len(self.days)):
            self.requestCheckbuttons.append([])
            self.requestVariables.append([])
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                     ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #a checkbutton should be active only if
                #the requested number of workers is greater than 0 for the given shift
                if self.cursor.fetchone()[0] > 0:
                    checkbutton = QtWidgets.QCheckBox()
                    self.workerRequestLayout.addWidget(checkbutton, 2+i, 1+j)
                    self.requestCheckbuttons[j].append(checkbutton)
                    #self.requestVariables[j].append(variable) #????
                else:
                    checkbutton = QtWidgets.QCheckBox()
                    checkbutton.setEnabled(False)
                    self.workerRequestLayout.addWidget(checkbutton, 2+i, 1+j)
                    self.requestCheckbuttons[j].append(checkbutton)
                    #self.requestVariables[j].append(variable) #????
        #print(self.requestVariables)

    def optionMenuSelectionEvent(self, event):
        '''
        event for selecting a name
        first it deselects all checkbuttons
        then it checks the shifts the worker requested for the given week
        '''
        for daysCheckbuttons in self.requestCheckbuttons:
            for checkbutton in daysCheckbuttons:
                checkbutton.setChecked(False)
        year = self.year
        week = self.week
        workerName = self.nameOptions.currentText()
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerIds = [row[0] for row in self.cursor.fetchall()]
                    if workerId in workerIds:
                        self.requestCheckbuttons[j][i].setChecked(True)
                except:
                    pass

    def getWorkerRequest(self):
        '''
        takes the checks from the check table into a numpy array (1 if checked, else 0)
        '''
        workerName = self.nameOptions.currentText()
        self.workerRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.workerRequestGrid[i][j] = 1 if self.requestCheckbuttons[j][i].isChecked() == True else 0 #when creating these checkbuttons and variables, the indices are reversed
        #print(workerName, '\n', self.workerRequestGrid)

    def saveWorkerRequest(self):
        '''
        saves worker requests for the given week to the database 
        '''
        self.getWorkerRequest()
        workerName = self.nameOptions.currentText()
        year = self.year
        week = self.week
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId, dayId, shiftId, UNIQUE(workerId, dayId, shiftId))')
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                if self.workerRequestGrid[i][j] == 1:
                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
                                        ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, j, i))
                else:
                    try:
                        self.cursor.execute('DELETE FROM workerRequests_' + str(year) + '_' + str(week) +
                                            ' WHERE workerId = ' + str(workerId) + ' AND dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    except:
                        pass
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Schedule manager

    def scheduleManager(self):
        '''
        gui for creating schedule
        '''
        tableExists = 1
        year = self.year
        week = self.week
        try:
            self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week))
        except:
            tableExists = 0
            
        if tableExists == 0:
            text = 'Table workerRequests_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
            self.messageWindow = QtWidgets.QWidget()
            layout = QtWidgets.QVBoxLayout() #layout
            label = QtWidgets.QLabel(text)
            layout.addWidget(label)
            self.messageWindow.setLayout(layout)
            self.messageWindow.show()
        else:
            self.scheduleWindow = QtWidgets.QWidget()
            self.scheduleWindow.setWindowTitle('Beosztás kezelése')
            self.layout = QtWidgets.QGridLayout()

            #header
            headerLabel = QtWidgets.QLabel('Beosztás kezelése')
            headerLayout = QtWidgets.QVBoxLayout()
            headerLayout.addWidget(headerLabel)
            self.layout.addLayout(headerLayout, 0, 0)

            #miscFrame
            yearLabel = QtWidgets.QLabel('Év')
            yearEntry = QtWidgets.QLineEdit()
            yearEntry.textChanged.connect(lambda x=yearEntry.text(): self.setYear(x))
            yearEntry.setText(str(self.year))
            weekLabel = QtWidgets.QLabel('Hét')
            weekEntry = QtWidgets.QLineEdit()
            weekEntry.textChanged.connect(lambda x=weekEntry.text(): self.setWeek(x))
            weekEntry.setText(str(self.week))
            showWorkerRequestsButton = QtWidgets.QPushButton('Ráérések kiírása')
            showWorkerRequestsButton.clicked.connect(self.showWorkerRequests)
            createScheduleButton = QtWidgets.QPushButton('Beosztás készítése')
            createScheduleButton.clicked.connect(self.createSchedule)
            fillCreatedScheduleButton = QtWidgets.QPushButton('Beosztás kiegészítése')
            fillCreatedScheduleButton.clicked.connect(self.fillCreatedSchedule)
            algorithmLabel = QtWidgets.QLabel('Algoritmus')
            self.algorithmOptions = QtWidgets.QComboBox()
            self.algorithmList = ['random', 'frommin']
            self.algorithmOptions.addItems(self.algorithmList)
            #self.algorithmOptions.activated.connect(self.nameMenuSelectionEvent)
            showScheduleButton = QtWidgets.QPushButton('Beosztás kiírása')
            showScheduleButton.clicked.connect(self.showSchedule)
            scheduleExportXlsxButton = QtWidgets.QPushButton('Export xlsx-be')
            scheduleExportXlsxButton.clicked.connect(self.scheduleExportXlsx)
            deleteScheduleButton = QtWidgets.QPushButton('Beosztás törlése')
            deleteScheduleButton.clicked.connect(self.deleteSchedule)
            miscLayout = QtWidgets.QGridLayout() #layout
            miscLayout.addWidget(yearLabel, 0, 0)
            miscLayout.addWidget(yearEntry, 0, 1)
            miscLayout.addWidget(weekLabel, 0, 2)
            miscLayout.addWidget(weekEntry, 0, 3)
            miscLayout.addWidget(showWorkerRequestsButton, 0, 4, 1, 2)
            miscLayout.addWidget(createScheduleButton, 1, 0, 1, 2)
            miscLayout.addWidget(fillCreatedScheduleButton, 1, 2, 1, 2)
            miscLayout.addWidget(algorithmLabel, 1, 4)
            miscLayout.addWidget(self.algorithmOptions, 1, 5)
            miscLayout.addWidget(showScheduleButton, 2, 0, 1, 2)
            miscLayout.addWidget(scheduleExportXlsxButton, 2, 2, 1, 2)
            miscLayout.addWidget(deleteScheduleButton, 3, 0, 1, 2)
            self.layout.addLayout(miscLayout, 1, 0)

            #scheduleFrame
            self.scheduleLayout = QtWidgets.QGridLayout() #layout
            self.layout.addLayout(self.scheduleLayout, 2, 0)

            self.scheduleWindow.setLayout(self.layout)
            self.scheduleWindow.show()

            self.showWorkerRequests()

    def loadRequestsListToShow(self, table):
        '''
        loads worker max number of request for shifts for the week into a list
        '''
        requests = list(range(len(self.shifts))) #[0]*len(self.shifts)
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                if table == 'workerRequests':
                    self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerIds = self.cursor.fetchall()
                    if len(workerIds) >= requests[i]:
                        requests[i] = len(workerIds)
                elif table == 'companyRequest':
                    self.cursor.execute('SELECT workerNumber FROM ' + table + '_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerNumber = self.cursor.fetchone()[0]
                    if workerNumber >= requests[i]:
                        requests[i] = workerNumber
        #print(table, requests)
        return requests

    def showWorkerRequests(self):
        '''
        creates a frame for handling worker requests for the given week
        checks if a worker is scheduled
        '''
        year = self.year
        week = self.week
        weekDay = self.weekDay
        row = 1 #same as gridRow
        requests = self.loadRequestsListToShow('workerRequests')
        #print('requests ready', requests)
        try:
            while self.scheduleLayout.count():
                widget = self.scheduleLayout.itemAt(0).widget()
                self.scheduleLayout.removeWidget(widget)
                widget.setParent(None)
        except:
            pass

        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        
        yearWeekLabel = QtWidgets.QLabel(str(year) + '/' + str(week))
        self.scheduleLayout.addWidget(yearWeekLabel, 0, 0)
        date = datetime.datetime.fromisocalendar(int(self.year), int(self.week), int(self.weekDay))
        startDate = date - datetime.timedelta(days=self.weekDay-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            label = QtWidgets.QLabel(text)
            self.scheduleLayout.addWidget(label, 0, 1+2*j, 1, 2)
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            self.scheduleLayout.addWidget(label, row, 0)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            self.scheduleByHandCheckbuttons.append([])
            self.scheduleByHandVariables.append([])
            self.scheduleByHandNameLabels.append([])
            gridRow = 1 #same as row
            gridRow_ = gridRow #to track the last empty row
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.scheduleByHandCheckbuttons[j].append([])
                self.scheduleByHandVariables[j].append([])
                self.scheduleByHandNameLabels[j].append([])
                label = QtWidgets.QLabel(self.shifts[i])
                self.scheduleLayout.addWidget(label, gridRow, 0)
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = self.cursor.fetchall()
                for k in range(0, requests[i]):
                    try:
                        workerId = workerIds[k][0]
                        self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + str(workerId))
                        workerName = self.cursor.fetchone()[0]
                        #nameLabel = QtWidgets.QLabel(workerName)
                        nameLabel = Label(workerName, self.scheduleLayout)
                        self.scheduleLayout.addWidget(nameLabel, gridRow_, 1+2*j)
                        self.scheduleByHandNameLabels[j][i].append(nameLabel)
                        checkbutton = QtWidgets.QCheckBox()
                        self.scheduleLayout.addWidget(checkbutton, gridRow_, 1+2*j+1)
                        self.scheduleByHandCheckbuttons[j][i].append(checkbutton)
                        self.scheduleByHandVariables[j][i].append([checkbutton, nameLabel])
                        try:
                            #check if the worker to be shown is already scheduled there (in a previous run of the program)
                            self.cursor.execute( 'SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                                 ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                            if workerId in [ workerIds[0] for workerIds in self.cursor.fetchall()]:
                                #if a worker is scheduled, check the box
                                checkbutton.setChecked(True)
                        except:
                            pass
                        checkbutton.stateChanged.connect(lambda x0, x1=j, x2=i, x3=k, x4=workerName: self.disableWorkerSelection(x0, x1, x2, x3, x4))
                    except:
                        #shitty solution to fill empty spaces
                        nameLabel = QtWidgets.QLabel('')
                        self.scheduleLayout.addWidget(nameLabel, gridRow_, 1+j)
                    gridRow_ += 1
                gridRow = gridRow + requests[i]
        #print(self.scheduleByHandVariables)

    def loadSchedule(self):
        '''
        loads the schedule and the backups for the given week
        '''
        year = self.year
        week = self.week
        self.schedule = [] #
        self.backup = [] #
        for j in range(0, len(self.days)):
            self.schedule.append([]) #
            self.backup.append([]) #
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.schedule[j].append(workerNames)
                
                #load the backup workers for the week (same as loading the scheduled workers)
                self.cursor.execute('SELECT workerId FROM backup_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.backup[j].append(workerNames)
        #print(self.schedule)

    def showSchedule(self):
        '''
        loads the schedule for the given week
        and shows it in a seperate window
        '''
        year = self.year
        week = self.week
        try:
            self.loadSchedule()
            self.showScheduleWindow = QtWidgets.QWidget()
            requests = self.loadRequestsListToShow('companyRequest')
            print(requests)
            row = 1 #starting row is the one under the day names

            #self.showScheduleNameLabels = []
            layout = QtWidgets.QGridLayout() #layout
            yearWeekLabel = QtWidgets.QLabel(str(year) + '/' + str(week))
            layout.addWidget(yearWeekLabel, 0, 0)
            for j in range(0, len(self.days)):
                label = QtWidgets.QLabel(self.days[j])
                layout.addWidget(label, 0, 1+j)
            for i in range(0, len(self.shifts)):
                label = QtWidgets.QLabel(self.shifts[i])
                layout.addWidget(label, row, 0)
                row = row + requests[i]
            for j in range(0, len(self.days)):
                #self.showScheduleNameLabels.append([])
                row_ = 1
                row = row_
                for i in range(0, len(self.shifts)):
                    for k in range(0, requests[i]):
                        try:
                            workerName = self.schedule[j][i][k]
                        except:
                            workerName = ''
                        #nameLabel = QtWidgets.QLabel(workerName)
                        nameLabel = Label(workerName, layout) #eddig ez a legjobb jelölt
                        layout.addWidget(nameLabel, row, 1+j)
                        #self.showScheduleNameLabels[j].append(nameLabel)
                        row += 1
                row_ = row_ + requests[i]
            
            self.showScheduleWindow.setLayout(layout)
            self.showScheduleWindow.show()

        except:
            text = 'Table schedule_' + str(year) + '_' + str(week) + ' does not exist.'
            print(text)
            self.showScheduleWindow = QtWidgets.QWidget()
            layout = QtWidgets.QVBoxLayout() #layout
            label = QtWidgets.QLabel(text)
            layout.addWidget(label)
            self.showScheduleWindow.setLayout(layout)
            self.showScheduleWindow.show()

    def scheduleExportXlsx(self):
        '''
        exports the schedule for the given week into a .xlsx file
        first loads the schedule from the database
        saves the backup workers for the week on a different worksheet (same as loading the scheduled workers)
        '''
        self.loadSchedule()
        year = self.year
        week = self.week
        weekDay = self.weekDay
        filename = 'schedule_' + str(year) + '_' + str(week) + '.xlsx'
        #schedule
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'schedule_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution below
        requests = self.loadRequestsListToShow('companyRequest')
        row = 2
        worksheet.cell(row=1, column=1).value = str(year) + '/' + str(week)
        worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        date = datetime.datetime.fromisocalendar(int(self.year), int(self.week), int(self.weekDay))
        startDate = date - datetime.timedelta(days=self.weekDay-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            worksheet.cell(row=1, column=2+j).value = text
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.schedule[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        #backup
        worksheet = workbook.create_sheet()
        worksheet.title = 'backup_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution above
        row = 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.backup[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        workbook.save(filename=filename)
        print('Schedule exported')

    def deleteSchedule(self):
        '''
        deletes schedule for the given week
        '''
        year = self.year
        week = self.week
        self.cursor.execute('DROP TABLE IF EXISTS schedule_' + str(year) + '_' + str(week))
        self.cursor.execute('DROP TABLE IF EXISTS backup_' + str(year) + '_' + str(week))
        self.saveDatabase()
    
    def disableWorkerSelection(self, CheckState, column, row, row_k, nameToDisable):
        '''
        if someone is scheduled to work in a shfit, he/she can't work in another shift on the given day
        the possibility to check him/her into another shift is disabled
        '''
        #print('disableWorkerSelection called')
        #print(self.scheduleByHandVariables[column][row][row_k][0].isChecked(), self.scheduleByHandVariables[column][row][row_k][1].text())
        #print(column, row, row_k, nameToDisable)
        
        if self.scheduleByHandVariables[column][row][row_k][0].isChecked() == True:
           for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k].text() == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k].setEnabled(False)
        else:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k].text() == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k].setEnabled(True)
        self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[column], ))
        dayId = self.cursor.fetchone()[0]
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[row], ))
        shiftId = self.cursor.fetchone()[0]
        self.disableWorkerSelectionForShift(dayId, shiftId, column, row)
        
    def disableWorkerSelectionForShift(self, dayId, shiftId, column, row):
        '''
        if the workers requested by the company for a given shift is met,
        the possibility to check other workers for that shift is disabled
        '''
        year = self.year
        week = self.week
        requests = self.loadRequestsListToShow('workerRequests') #gives the max number of requests for shifts
        workersScheduledForShift = []
        workerNumberScheduled = 0
        workersScheduledForDay = []
        for k in range(0, requests[row]):
            #try-except is not the most elegant solution
            #it is for overcoming that requests list contains the max number of requests for shifts (for example [8, 1, 5])
            #and the real requests for a given day can be fewer (for example [6, 1, 4])
            #so the index k may result in out of range error
            #and company requests is [4, 1, 4]
            try:
                if self.scheduleByHandVariables[column][row][k][0].isChecked() == True:
                    workerNumberScheduled += 1
                    workersScheduledForShift.append(self.scheduleByHandNameLabels[column][row][k].text())
            except:
                pass
        #print('workersScheduledForShift: ', workersScheduledForShift)

        for row_ in range(0, len(requests)):
            for k in range(0, requests[row_]):
                try:
                    if self.scheduleByHandVariables[column][row_][k][0].isChecked() == True:
                        workersScheduledForDay.append(self.scheduleByHandNameLabels[column][row_][k].text())
                except:
                    pass
        #print('workersScheduledForDay: ', workersScheduledForDay)

        self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) + 
                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
        workerNumber = self.cursor.fetchone()[0]
        if workerNumberScheduled == workerNumber:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k].text()
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k].setEnabled(False)
            except:
                pass
        else:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k].text()
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k].setEnabled(True)
            except:
                pass
            
    def createBackup(self):
        '''
        creates a backup table for the given week
        from the workers who are not scheduled
        '''
        year = self.year
        week = self.week
        self.cursor.execute('DROP TABLE IF EXISTS backup_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE backup_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) )
                workerIdsSchedule = self.cursor.fetchall()
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIdsRequests = self.cursor.fetchall()
                for id_ in workerIdsRequests:
                    if not id_ in workerIdsSchedule:
                        row = id_[0], dayId, shiftId
                        self.cursor.execute('INSERT OR IGNORE INTO backup_' + str(year) + '_' + str(week) +
                                            ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', row)
        print('Backup created')

    def createSchedule(self):
        '''
        creates schedule from the check table
        also calls createBackup()
        '''
        year = self.year
        week = self.week
        self.cursor.execute('DROP TABLE IF EXISTS schedule_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE schedule_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for day in range(0, len(self.scheduleByHandCheckbuttons)):
            for shift in range(0, len(self.scheduleByHandCheckbuttons[day])):
                for i in range(0, len(self.scheduleByHandCheckbuttons[day][shift])):
                    checkbutton = self.scheduleByHandCheckbuttons[day][shift][i]
                    if checkbutton.isChecked() == True:
                        workerName = self.scheduleByHandNameLabels[day][shift][i].text()
                        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName, ))
                        workerId = self.cursor.fetchone()[0]
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, day, shift) )
        self.createBackup()
        self.saveDatabase()
        print('Schedule created')


    def fillCreatedSchedule(self):
        '''
        completes and saves the schedule from the check table based on the selected algorithm
        also calls createBackup() and then getNumberOfScheduledDays()
        '''
        year = self.year
        week = self.week
        self.createSchedule()
        self.cursor.execute('DROP TABLE IF EXISTS companyRequestModified') #this modified table is for counting how many workers are still needed 
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequestModified AS SELECT * FROM companyRequest_'
                            + str(year) + '_' + str(week) +' WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                partialScheduledWorkers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNeeded = self.cursor.fetchone()[0]
                if workerNeeded > len(partialScheduledWorkers):
                    workerNeeded -= len(partialScheduledWorkers)
                else:
                    workerNeeded = 0
                self.cursor.execute('INSERT INTO companyRequestModified (dayId, shiftId, workerNumber) VALUES (?, ?, ?) ', (dayId, shiftId, workerNeeded))

        if self.algorithmOptions.currentText() == 'random':
            #completes the schedule randomly, may not schedule workers who are free anyway
            self.cursor.execute( 'SELECT * FROM workerRequests_' + str(year) + '_' + str(week) )
            workerRequests = self.cursor.fetchall()
            random.shuffle(workerRequests)
            for row in workerRequests:
                workerId, dayId, shiftId = row[0], row[1], row[2]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + #select workerId instead of *
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workersNeeded = self.cursor.fetchone()[0]
                if len(workers) < workersNeeded:
                    if not workerId in workers:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        elif self.algorithmOptions.currentText() == 'frommin':
            #completes the table starting from the workers who requested the least days
            self.getNumberOfRequestedDays()
            self.cursor.execute('SELECT * FROM workers ORDER BY requestedDaysWeekly')
            workers = self.cursor.fetchall()
            for worker in workers:
                workerId = worker[0]
                self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE workerId = ' + str(workerId) )
                for row in self.cursor.fetchall():
                    dayId, shiftId = row[1], row[2]
                    self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + 
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workers = self.cursor.fetchall()
                    self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workersNeeded = self.cursor.fetchone()[0]
                    if len(workers) < workersNeeded:
                        if not workerId in workers:
                            self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                                '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        self.createBackup()
        self.getNumberOfScheduledDays()
        self.saveDatabase()
        print('Schedule created and filled')

    def getNumberOfRequestedDays(self):
        '''
        determines how many days each worker has requested for the week
        '''
        self.numberOfRequestedDays = {}
        year = self.year
        week = self.week
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM workerRequests_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfRequestedDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET requestedDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfRequestedDays:', self.numberOfRequestedDays)

    def getNumberOfScheduledDays(self):
        '''
        determines how many days each worker has been scheduled for for the week
        '''
        self.numberOfScheduledDays = {}
        year = self.year
        week = self.week
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM schedule_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfScheduledDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET scheduledDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfScheduledDays: ' + self.numberOfScheduledDays)


        
if __name__ == '__main__':
    app = SHScheduler(sys.argv)
    app.setStyle('Fusion')
    app.exec()
